import os

from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from employee_manager import EmployeeManager
from health_calculator import HealthCalculator
from fuzzy_system import FuzzyRiskSystem

from normalizers import ParameterNormalizer

DB_URL = 'database/risk_assesment.db'

class ConfigInfoDialog(QDialog):
    """Диалог отображения информации о конфигурации"""

    def __init__(self, parent=None, config_info=""):
        super().__init__(parent)
        self.setup_ui(config_info)

    def setup_ui(self, config_info):
        self.setWindowTitle("Информация о конфигурации")
        self.setMinimumSize(600, 400)

        layout = QVBoxLayout()

        # Текстовое поле с информацией
        self.info_text = QTextEdit()
        self.info_text.setReadOnly(True)
        self.info_text.setFont(QFont("Courier New", 10))
        self.info_text.setText(config_info)

        layout.addWidget(self.info_text)

        # Кнопка закрытия
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        self.setLayout(layout)


class EmployeeTableModel(QAbstractTableModel):
    """Модель таблицы для отображения работников"""

    def __init__(self, employees):
        super().__init__()
        self.employees = employees
        self.headers = ['№', 'ФИО', 'Должность', 'Предприятие', 'Пол', 'Возраст', 'Дата приема на работу', 'Проф. вредность', 'Год вредности', 'Инвалидность',
                        'Диагнозы', 'Здоровье']

    def rowCount(self, parent=None):
        return len(self.employees)

    def columnCount(self, parent=None):
        return len(self.headers)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None

        employee = self.employees[index.row()]
        col = index.column()

        if role == Qt.ItemDataRole.DisplayRole:
            try:
                if col == 0:
                    return str(employee.id)
                elif col == 1:
                    return employee.full_name
                elif col == 2:
                    return employee.position
                elif col == 3:
                    return employee.department_name
                elif col == 4:
                    return employee.gender
                elif col == 5:
                    age = employee.get_age()
                    return str(age) if age is not None else ""
                elif col == 6:
                    return employee.start_year if hasattr(employee, 'start_year') and employee.start_year else ""
                elif col == 11:
                    score = HealthCalculator.calculate_health_score(employee)
                    #desc = HealthCalculator.get_health_description(score)
                    return f"{score:.2f}"
                    #return f"{score:.2f} ({desc})"
                elif col == 10:
                    if hasattr(employee, 'diagnoses') and employee.diagnoses:
                        all_diagnoses = []
                        for category, diagnoses in employee.diagnoses.items():
                            if diagnoses:
                                all_diagnoses.extend(diagnoses)  # По одному из каждой категории
                        return ", ".join(all_diagnoses)

                    return ""
                elif col == 7:
                    return employee.prof_harm_code if hasattr(employee,
                                                              'prof_harm_code') and employee.prof_harm_code else ""
                elif col == 9:
                    if hasattr(employee, 'disability_group') and employee.disability_group:
                        return f"Гр.{employee.disability_group}"
                    return ""
                elif col == 8:
                    return employee.prof_harm_year if hasattr(employee,
                                                              'prof_harm_year') and employee.prof_harm_year else ""

            except Exception as e:
                print(f"Error getting data for cell {index.row()}, {col}: {e}")
                return "Ошибка"

        elif role == Qt.ItemDataRole.TextAlignmentRole:
            if col in [0, 3, 4, 5, 6, 7, 8, 9, 11]:
                return Qt.AlignmentFlag.AlignCenter
            return Qt.AlignmentFlag.AlignLeft

        elif role == Qt.ItemDataRole.ToolTipRole:
            if col == 10 and hasattr(employee, 'diagnoses') and employee.diagnoses:
                tooltip = "<b>Диагнозы:</b><br>"
                for category, diagnoses in employee.diagnoses.items():
                    if diagnoses:
                        tooltip += f"<b>{category}:</b><br>"
                        for diag in diagnoses:
                            tooltip += f"• {diag}<br>"
                        tooltip += "<br>"
                return tooltip

        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if orientation == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
            return self.headers[section]
        return None


class DiagnosisInputWidget(QWidget):
    """Виджет для ввода диагнозов"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.diagnosis_widgets = []  # Список виджетов для каждого диагноза
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Заголовок
        title_label = QLabel("Диагнозы:")
        title_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(title_label)

        # Контейнер для виджетов диагнозов
        self.diagnoses_container = QVBoxLayout()
        layout.addLayout(self.diagnoses_container)

        # Кнопка добавления нового диагноза
        add_btn = QPushButton("+ Добавить диагноз")
        add_btn.clicked.connect(self.add_diagnosis_widget)
        layout.addWidget(add_btn)

    def add_diagnosis_widget(self):
        """Добавить виджет для ввода одного диагноза"""
        diagnosis_widget = QWidget()
        diagnosis_layout = QHBoxLayout(diagnosis_widget)

        # Выбор категории
        category_combo = QComboBox()
        category_combo.addItem("Выберите категорию...", "")

        # Загружаем категории из БД
        from employee_manager import EmployeeManager
        manager = EmployeeManager(DB_URL)
        categories = manager.get_diagnosis_categories()

        for cat in categories:
            category_combo.addItem(cat['category'], cat['category'])

        # Поле для ввода названия диагноза
        diagnosis_edit = QLineEdit()
        diagnosis_edit.setPlaceholderText("Введите название диагноза...")

        # Кнопка удаления
        remove_btn = QPushButton("✕")
        remove_btn.setMaximumWidth(30)
        remove_btn.clicked.connect(lambda: self.remove_diagnosis_widget(diagnosis_widget))

        diagnosis_layout.addWidget(category_combo)
        diagnosis_layout.addWidget(diagnosis_edit, 1)  # Растягиваемое поле
        diagnosis_layout.addWidget(remove_btn)

        # Сохраняем ссылки на виджеты
        diagnosis_data = {
            'widget': diagnosis_widget,
            'category_combo': category_combo,
            'diagnosis_edit': diagnosis_edit
        }
        self.diagnosis_widgets.append(diagnosis_data)

        self.diagnoses_container.addWidget(diagnosis_widget)

    def remove_diagnosis_widget(self, widget):
        """Удалить виджет диагноза"""
        for i, data in enumerate(self.diagnosis_widgets):
            if data['widget'] == widget:
                widget.setParent(None)
                widget.deleteLater()
                self.diagnosis_widgets.pop(i)
                break

    def get_diagnoses(self):
        """Получить все введенные диагнозы"""
        diagnoses_by_category = {}

        for data in self.diagnosis_widgets:
            category = data['category_combo'].currentData()
            diagnosis_name = data['diagnosis_edit'].text().strip()

            if category and diagnosis_name:
                if category not in diagnoses_by_category:
                    diagnoses_by_category[category] = []

                if diagnosis_name not in diagnoses_by_category[category]:
                    diagnoses_by_category[category].append(diagnosis_name)

        return diagnoses_by_category

    def set_diagnoses(self, diagnoses_dict):
        """Установить диагнозы из существующего сотрудника"""
        # Удаляем все текущие виджеты
        for data in self.diagnosis_widgets[:]:
            self.remove_diagnosis_widget(data['widget'])

        # Добавляем виджеты для каждого диагноза
        if diagnoses_dict:
            for category, diagnosis_list in diagnoses_dict.items():
                for diagnosis_name in diagnosis_list:
                    self.add_diagnosis_widget()

                    # Устанавливаем значения в последнем добавленном виджете
                    if self.diagnosis_widgets:
                        last_data = self.diagnosis_widgets[-1]

                        # Устанавливаем категорию
                        index = last_data['category_combo'].findData(category)
                        if index >= 0:
                            last_data['category_combo'].setCurrentIndex(index)

                        # Устанавливаем название диагноза
                        last_data['diagnosis_edit'].setText(diagnosis_name)


class AddEditEmployeeDialog(QDialog):
    """Диалог добавления/редактирования работника"""

    def __init__(self, parent=None, employee=None):
        super().__init__(parent)
        self.employee = employee
        self.setup_ui()

        if employee:
            self.load_employee_data()

    def setup_ui(self):
        self.setWindowTitle("Добавить работника" if not self.employee else "Редактировать работника")
        self.setMinimumWidth(600)

        # Основной layout
        main_layout = QVBoxLayout()

        # Scroll area для прокрутки
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)

        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)

        # 1. Основная информация
        info_group = QGroupBox("Основная информация")
        info_layout = QFormLayout()

        self.lastname_edit = QLineEdit()
        info_layout.addRow("Фамилия:", self.lastname_edit)

        self.firstname_edit = QLineEdit()
        info_layout.addRow("Имя:", self.firstname_edit)

        self.patronymic_edit = QLineEdit()
        self.patronymic_edit.setPlaceholderText("необязательно")
        info_layout.addRow("Отчество:", self.patronymic_edit)

        # Должность
        self.position_combo = QComboBox()
        from employee_manager import EmployeeManager
        manager = EmployeeManager(DB_URL)
        positions = manager.get_positions()
        for pos in positions:
            self.position_combo.addItem(pos['name'], pos['id'])
        info_layout.addRow("Должность:", self.position_combo)

        # Предприятие
        self.department_combo = QComboBox()
        departments = manager.get_departments()
        for dept in departments:
            self.department_combo.addItem(dept['name'], dept['id'])
        info_layout.addRow("Предприятие:", self.department_combo)

        # Пол
        self.gender_combo = QComboBox()
        self.gender_combo.addItems(["М", "Ж"])
        info_layout.addRow("Пол:", self.gender_combo)

        # Дата рождения
        self.birth_date_edit = QDateEdit()
        self.birth_date_edit.setCalendarPopup(True)
        self.birth_date_edit.setDate(QDate.currentDate().addYears(-30))
        info_layout.addRow("Дата рождения:", self.birth_date_edit)

        # Дата приема на работу
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(QDate.currentDate())
        info_layout.addRow("Дата приема на работу:", self.start_date_edit)

        info_group.setLayout(info_layout)
        content_layout.addWidget(info_group)

        # Диагнозы
        self.diagnosis_widget = DiagnosisInputWidget()
        content_layout.addWidget(self.diagnosis_widget)

        # Профвредность
        prof_group = QGroupBox("Профессиональная вредность")
        prof_layout = QFormLayout()

        self.prof_harm_edit = QLineEdit()
        self.prof_harm_edit.setPlaceholderText("Например: Т75.2")
        prof_layout.addRow("Код профвредности:", self.prof_harm_edit)

        self.prof_harm_year_edit = QDateEdit()
        self.prof_harm_year_edit.setCalendarPopup(True)
        self.prof_harm_year_edit.setDate(QDate.currentDate())
        prof_layout.addRow("Год установления:", self.prof_harm_year_edit)

        prof_group.setLayout(prof_layout)
        content_layout.addWidget(prof_group)

        # Инвалидность
        disability_group = QGroupBox("Инвалидность")
        disability_layout = QFormLayout()

        self.disability_combo = QComboBox()
        self.disability_combo.addItem("Нет", None)
        self.disability_combo.addItem("1 группа", 1)
        self.disability_combo.addItem("2 группа", 2)
        self.disability_combo.addItem("3 группа", 3)
        disability_layout.addRow("Группа инвалидности:", self.disability_combo)

        disability_group.setLayout(disability_layout)
        content_layout.addWidget(disability_group)

        # Добавляем растягивающийся спейсер
        content_layout.addStretch()

        # Устанавливаем виджет в scroll area
        scroll.setWidget(content_widget)
        main_layout.addWidget(scroll)

        # Кнопки
        button_layout = QHBoxLayout()
        self.save_btn = QPushButton("Сохранить")
        self.cancel_btn = QPushButton("Отмена")

        self.save_btn.clicked.connect(self.validate_and_accept)
        self.cancel_btn.clicked.connect(self.reject)

        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.cancel_btn)
        main_layout.addLayout(button_layout)

        self.setLayout(main_layout)

    def load_employee_data(self):
        """Загрузка данных работника в форму"""
        if not self.employee:
            return

        self.lastname_edit.setText(self.employee.lastname)
        if hasattr(self.employee, 'firstname'):
            self.firstname_edit.setText(self.employee.firstname or '')
        if hasattr(self.employee, 'patronymic'):
            self.patronymic_edit.setText(self.employee.patronymic or '')

        # Установка должности
        for i in range(self.position_combo.count()):
            if self.position_combo.itemText(i) == self.employee.position:
                self.position_combo.setCurrentIndex(i)
                break

        # Установка предприятия
        if hasattr(self.employee, 'department_id') and self.employee.department_id:
            for i in range(self.department_combo.count()):
                if self.department_combo.itemData(i) == self.employee.department_id:
                    self.department_combo.setCurrentIndex(i)
                    break

        # Установка пола
        index = self.gender_combo.findText(self.employee.gender)
        if index >= 0:
            self.gender_combo.setCurrentIndex(index)

        # Даты
        if hasattr(self.employee, 'birth_date') and self.employee.birth_date:
            try:
                birth_date = QDate.fromString(str(self.employee.birth_date), 'yyyy-MM-dd')
                self.birth_date_edit.setDate(birth_date)
            except:
                pass

        if hasattr(self.employee, 'start_year') and self.employee.start_year:
            try:
                start_date = QDate.fromString(str(self.employee.start_year), 'yyyy-MM-dd')
                self.start_date_edit.setDate(start_date)
            except:
                pass

        # Диагнозы
        if hasattr(self.employee, 'diagnoses') and self.employee.diagnoses:
            self.diagnosis_widget.set_diagnoses(self.employee.diagnoses)

        # Профвредность
        if hasattr(self.employee, 'prof_harm_code') and self.employee.prof_harm_code:
            self.prof_harm_edit.setText(self.employee.prof_harm_code)

        if hasattr(self.employee, 'prof_harm_year') and self.employee.prof_harm_year:
            try:
                prof_year = QDate.fromString(str(self.employee.prof_harm_year), 'yyyy-MM-dd')
                self.prof_harm_year_edit.setDate(prof_year)
            except:
                pass

        # Инвалидность
        if hasattr(self.employee, 'disability_group') and self.employee.disability_group:
            for i in range(self.disability_combo.count()):
                if self.disability_combo.itemData(i) == self.employee.disability_group:
                    self.disability_combo.setCurrentIndex(i)
                    break

    def validate_and_accept(self):
        """Проверка данных и закрытие диалога"""
        # Проверяем обязательные поля
        if not self.lastname_edit.text().strip():
            QMessageBox.warning(self, "Ошибка", "Введите фамилию работника!")
            return

        # Получаем диагнозы
        diagnoses = self.diagnosis_widget.get_diagnoses()

        # Формируем данные
        employee_data = {
            'lastname': self.lastname_edit.text().strip(),
            'firstname': self.firstname_edit.text().strip(),
            'patronymic': self.patronymic_edit.text().strip(),
            'position_id': self.position_combo.currentData(),
            'gender': self.gender_combo.currentText(),
            'birth_date': self.birth_date_edit.date().toString('yyyy-MM-dd'),
            'start_year': self.start_date_edit.date().toString('yyyy-MM-dd'),
            'department_id': self.department_combo.currentData(),
            'diagnoses': diagnoses
        }

        # Профвредность
        prof_harm_code = self.prof_harm_edit.text().strip()
        if prof_harm_code:
            employee_data['prof_harm_code'] = prof_harm_code
            employee_data['prof_harm_year'] = self.prof_harm_year_edit.date().toString('yyyy-MM-dd')

        # Инвалидность
        disability_group = self.disability_combo.currentData()
        if disability_group is not None:
            employee_data['disability_group'] = disability_group

        # Сохраняем данные
        self.employee_data = employee_data
        self.accept()

    def get_employee_data(self):
        """Получить данные из формы"""
        return getattr(self, 'employee_data', None)


class ParameterInputWidget(QWidget):
    """Виджет для ввода физического параметра с нормализацией"""

    def __init__(self, param_name, param_type, parent=None):
        super().__init__(parent)
        self.param_name = param_name
        self.param_type = param_type
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(5)

        # Заголовок с единицами измерения
        info = ParameterNormalizer.get_parameter_info(self.param_type)
        self.title_label = QLabel(f"{self.param_name} ({info['description']})")
        self.title_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(self.title_label)

        # Слайдер (в нормализованной шкале 0-100)
        slider_layout = QHBoxLayout()
        self.slider = QSlider(Qt.Orientation.Horizontal)
        self.slider.setRange(0, 100)
        self.slider.setValue(0)
        slider_layout.addWidget(self.slider)

        # Отображение физического значения
        self.value_label = QLabel("0.00")
        self.value_label.setMinimumWidth(80)
        self.value_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        slider_layout.addWidget(self.value_label)

        layout.addLayout(slider_layout)

        # Отображение нормализованного значения
        norm_layout = QHBoxLayout()
        norm_layout.addWidget(QLabel("Нормализованное значение:"))
        self.norm_label = QLabel("0.00")
        self.norm_label.setStyleSheet("color: #666;")
        norm_layout.addWidget(self.norm_label)
        norm_layout.addStretch()

        layout.addLayout(norm_layout)

        # Подключение сигналов
        self.slider.valueChanged.connect(self.on_slider_changed)

        self.setLayout(layout)

        # Текущее физическое значение
        self.physical_value = 0.0

    def on_slider_changed(self, norm_value):
        """При изменении слайдера обновляем физическое значение"""
        normalized = norm_value / 100.0

        # Конвертируем в физическое значение
        if self.param_type == 'noise':
            physical = ParameterNormalizer.denormalize_noise(normalized)
        elif self.param_type == 'vibration':
            physical = ParameterNormalizer.denormalize_vibration(normalized)
        elif self.param_type == 'chemical':
            physical = ParameterNormalizer.denormalize_chemical(normalized)
        else:
            physical = normalized

        self.physical_value = physical

        # Обновляем отображение
        if self.param_type == 'chemical':
            self.value_label.setText(f"{physical:.3f}")
        else:
            self.value_label.setText(f"{physical:.1f}")

        self.norm_label.setText(f"{normalized:.2f}")

    def set_physical_value(self, physical_value):
        """Установить физическое значение (обновляет слайдер)"""
        self.physical_value = physical_value

        # Конвертируем в нормализованное
        if self.param_type == 'noise':
            normalized = ParameterNormalizer.normalize_noise(physical_value)
        elif self.param_type == 'vibration':
            normalized = ParameterNormalizer.normalize_vibration(physical_value)
        elif self.param_type == 'chemical':
            normalized = ParameterNormalizer.normalize_chemical(physical_value)
        else:
            normalized = physical_value

        # Обновляем слайдер без генерации сигнала
        self.slider.blockSignals(True)
        self.slider.setValue(int(normalized * 100))
        self.slider.blockSignals(False)

        # Обновляем отображение
        if self.param_type == 'chemical':
            self.value_label.setText(f"{physical_value:.3f}")
        else:
            self.value_label.setText(f"{physical_value:.1f}")

        self.norm_label.setText(f"{normalized:.2f}")

    def get_normalized_value(self):
        """Получить нормализованное значение (0-1)"""
        return self.slider.value() / 100.0

    def get_physical_value(self):
        """Получить физическое значение"""
        return self.physical_value


class MultiRiskCalculatorDialog(QDialog):
    """Диалог расчета риска для нескольких сотрудников"""

    def __init__(self, parent=None, employees=None, fuzzy_system=None):
        super().__init__(parent)
        self.employees = employees if employees else []
        self.results = {}

        if fuzzy_system:
            self.fuzzy_system = fuzzy_system
        else:
            self.fuzzy_system = FuzzyRiskSystem()

        self.setup_ui()
        self.load_employees_data()

    def setup_ui(self):
        self.setWindowTitle("Расчет риска для сотрудников")
        self.resize(1000, 750)

        layout = QVBoxLayout()

        # 1. Панель с входными параметрами
        params_group = QGroupBox("Параметры рабочей среды")
        params_layout = QVBoxLayout()

        # Создаем виджеты для каждого параметра
        self.noise_widget = ParameterInputWidget("Шум", "noise")
        self.vibration_widget = ParameterInputWidget("Вибрация", "vibration")
        self.chemical_widget = ParameterInputWidget("Химический фактор", "chemical")

        params_layout.addWidget(self.noise_widget)
        params_layout.addWidget(self.vibration_widget)
        params_layout.addWidget(self.chemical_widget)


        params_group.setLayout(params_layout)
        layout.addWidget(params_group)

        # 2. Кнопка расчета для всех
        calc_all_layout = QHBoxLayout()
        self.calculate_all_btn = QPushButton("Рассчитать риск для всех")
        self.calculate_all_btn.clicked.connect(self.calculate_risk_for_all)
        self.calculate_all_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        calc_all_layout.addWidget(self.calculate_all_btn)
        layout.addLayout(calc_all_layout)

        # 3. Таблица с сотрудниками и результатами
        table_group = QGroupBox("Результаты расчета")
        table_layout = QVBoxLayout()

        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(6)
        self.table_widget.setHorizontalHeaderLabels([
            "ФИО", "Должность", "Предприятие",
            "Показатель здоровья", "Уровень риска", "Категория риска"
        ])

        # Настройка таблицы
        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)

        self.table_widget.setAlternatingRowColors(True)

        table_layout.addWidget(self.table_widget)
        table_group.setLayout(table_layout)
        layout.addWidget(table_group)

        # 4. Статистика
        stats_layout = QHBoxLayout()
        self.stats_label = QLabel("Выбрано сотрудников: 0")
        stats_layout.addWidget(self.stats_label)
        stats_layout.addStretch()

        # Индикатор текущих параметров
        self.params_info = QLabel("")
        self.params_info.setStyleSheet("color: #666; font-size: 10px;")
        stats_layout.addWidget(self.params_info)

        layout.addLayout(stats_layout)

        # 5. Кнопки
        button_layout = QHBoxLayout()

        self.export_results_btn = QPushButton("Экспорт результатов")
        self.export_results_btn.clicked.connect(self.export_results)

        self.close_btn = QPushButton("Закрыть")
        self.close_btn.clicked.connect(self.accept)

        button_layout.addWidget(self.export_results_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.close_btn)

        layout.addLayout(button_layout)

        self.setLayout(layout)

        self.update_params_info()

    def update_params_info(self):
        """Обновить информацию о текущих параметрах"""
        noise_val = self.noise_widget.get_physical_value()
        vib_val = self.vibration_widget.get_physical_value()
        chem_val = self.chemical_widget.get_physical_value()

    def load_employees_data(self):
        """Загрузка данных сотрудников в таблицу"""
        self.table_widget.setRowCount(len(self.employees))

        for row, employee in enumerate(self.employees):
            # ФИО
            name_item = QTableWidgetItem(employee.full_name)
            name_item.setData(Qt.ItemDataRole.UserRole, employee.id)
            self.table_widget.setItem(row, 0, name_item)

            # Должность
            self.table_widget.setItem(row, 1, QTableWidgetItem(employee.position))

            # Предприятие
            dept_item = QTableWidgetItem(str(employee.department_id))
            dept_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table_widget.setItem(row, 2, dept_item)

            # Показатель здоровья (используем новый метод health_deficit)
            health_deficit = HealthCalculator.calculate_health_score(employee)
            health_item = QTableWidgetItem(f"{health_deficit:.3f}")
            health_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            self.table_widget.setItem(row, 3, health_item)

            # Уровень риска (пока пусто)
            risk_item = QTableWidgetItem("-")
            risk_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table_widget.setItem(row, 4, risk_item)

            # Категория риска (пока пусто)
            category_item = QTableWidgetItem("-")
            category_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table_widget.setItem(row, 5, category_item)

        self.stats_label.setText(f"Выбрано сотрудников: {len(self.employees)}")

    def calculate_risk_for_all(self):
        """Рассчитать риск для всех сотрудников"""
        try:
            # Получаем нормализованные значения параметров
            vibration_val = self.vibration_widget.get_normalized_value()
            noise_val = self.noise_widget.get_normalized_value()
            chemical_val = self.chemical_widget.get_normalized_value()

            # Валидация
            vibration_val = max(0.0, min(1.0, vibration_val))
            noise_val = max(0.0, min(1.0, noise_val))
            chemical_val = max(0.0, min(1.0, chemical_val))

            # Обновляем информацию
            self.update_params_info()

            # Рассчитываем риск для каждого сотрудника
            for row, employee in enumerate(self.employees):
                health_val = HealthCalculator.calculate_health_score(employee)

                result = self.fuzzy_system.calculate_risk(
                    vibration_val, noise_val, chemical_val, health_val
                )

                if result['success']:
                    # Сохраняем результат
                    self.results[employee.id] = result

                    # Обновляем таблицу
                    risk_item = QTableWidgetItem(f"{result['value']:.3f}")
                    risk_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                    self.table_widget.setItem(row, 4, risk_item)

                    category_item = QTableWidgetItem(result['category'])
                    category_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table_widget.setItem(row, 5, category_item)

            self.calculate_all_btn.setStyleSheet("background-color: #45a049; color: white; font-weight: bold;")
            self.statusBar().showMessage("Расчет выполнен", 3000) if hasattr(self, 'statusBar') else None

        except ValueError as e:
            QMessageBox.warning(self, "Ошибка ввода",
                                f"Пожалуйста, введите корректные значения:\n{str(e)}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка расчета",
                                 f"Произошла ошибка при расчете: {str(e)}")

    def export_results(self):
        """Экспорт результатов в Excel"""
        if not self.results:
            QMessageBox.warning(self, "Ошибка", "Нет результатов для экспорта!")
            return

        try:
            import pandas as pd
            from datetime import datetime

            data = []
            for row, employee in enumerate(self.employees):
                if employee.id in self.results:
                    result = self.results[employee.id]
                    health_deficit = HealthCalculator.calculate_health_score(employee)

                    data.append({
                        'ФИО': employee.full_name,
                        'Должность': employee.position,
                        'Предприятие': employee.department_id,
                        'Показатель здоровья (дефицит)': f"{health_deficit:.3f}",
                        'Уровень риска': f"{result['value']:.3f}",
                        'Категория риска': result['category'],
                        'Вибрация (дБ)': f"{self.vibration_widget.get_physical_value():.1f}",
                        'Шум (дБА)': f"{self.noise_widget.get_physical_value():.1f}",
                        'Марганец (мг/м³)': f"{self.chemical_widget.get_physical_value():.3f}",
                        'Вибрация (норм.)': f"{self.vibration_widget.get_normalized_value():.3f}",
                        'Шум (норм.)': f"{self.noise_widget.get_normalized_value():.3f}",
                        'Марганец (норм.)': f"{self.chemical_widget.get_normalized_value():.3f}"
                    })

            df = pd.DataFrame(data)

            filename = f"risk_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить результаты",
                filename,
                "Excel Files (*.xlsx);;All Files (*)"
            )

            if file_path:
                df.to_excel(file_path, index=False)
                QMessageBox.information(self, "Успешно",
                                        f"Результаты экспортированы в файл:\n{file_path}")

        except ImportError:
            QMessageBox.warning(self, "Ошибка",
                                "Для экспорта результатов необходима библиотека pandas")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать: {str(e)}")

class MainWindow(QMainWindow):
    """Главное окно приложения"""

    def __init__(self):
        super().__init__()

        # Инициализируем менеджер сотрудников
        self.employee_manager = EmployeeManager(DB_URL)

        # Инициализируем систему нечеткой логики
        try:
            from fuzzy_system import FuzzyRiskSystem
            self.fuzzy_system = FuzzyRiskSystem()
        except ImportError:
            print("Warning: Fuzzy system not available")
            self.fuzzy_system = None

        self.current_config_file = None
        self.setup_ui()
        self.load_employees()

    def setup_ui(self):
        self.setWindowTitle("Система оценки рисков здоровья работников")
        self.resize(1000, 600)

        # Центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Основной layout
        main_layout = QVBoxLayout(central_widget)

        self.create_menu_bar()

        # Панель поиска с тремя полями
        search_layout = QHBoxLayout()

        # Поле поиска по ФИО
        self.search_name_edit = QLineEdit()
        self.search_name_edit.setPlaceholderText("Поиск по фамилии, имени, отчеству...")
        self.search_name_edit.textChanged.connect(self.search_employees)
        search_layout.addWidget(self.search_name_edit)

        # Поле поиска по должности
        self.search_position_edit = QLineEdit()
        self.search_position_edit.setPlaceholderText("Поиск по должности...")
        self.search_position_edit.textChanged.connect(self.search_employees)
        search_layout.addWidget(self.search_position_edit)

        # Поле поиска по предприятию
        self.search_department_edit = QLineEdit()
        self.search_department_edit.setPlaceholderText("Поиск по номеру предприятия...")
        self.search_department_edit.textChanged.connect(self.search_employees)
        search_layout.addWidget(self.search_department_edit)

        # Кнопка сброса поиска
        self.clear_search_btn = QPushButton("Сбросить")
        self.clear_search_btn.clicked.connect(self.clear_search)
        search_layout.addWidget(self.clear_search_btn)

        main_layout.addLayout(search_layout)

        # Панель выбора сотрудников
        selection_layout = QHBoxLayout()

        self.select_all_btn = QPushButton("Выбрать всех")
        self.select_all_btn.clicked.connect(self.select_all_employees)

        self.deselect_all_btn = QPushButton("Снять выделение")
        self.deselect_all_btn.clicked.connect(self.deselect_all_employees)

        selection_layout.addWidget(self.select_all_btn)
        selection_layout.addWidget(self.deselect_all_btn)
        selection_layout.addStretch()

        main_layout.addLayout(selection_layout)

        # Таблица работников
        self.table_view = QTableView()
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_view.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        self.table_view.doubleClicked.connect(self.on_table_double_click)

        # Настройки таблицы
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setShowGrid(True)  # Показывать сетку
        #self.table_view.setSortingEnabled(True)

        #self.table_view.horizontalHeader().setStretchLastSection(True)
        #self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        main_layout.addWidget(self.table_view)

        # Панель кнопок
        button_layout = QHBoxLayout()

        self.add_btn = QPushButton("Добавить работника")
        self.add_btn.clicked.connect(self.add_employee)

        self.edit_btn = QPushButton("Редактировать")
        self.edit_btn.clicked.connect(self.edit_selected_employee)

        self.delete_btn = QPushButton("Удалить")
        self.delete_btn.clicked.connect(self.delete_selected_employee)

        self.calculate_btn = QPushButton("Рассчитать риск")
        self.calculate_btn.clicked.connect(self.calculate_risk_for_selected)
        self.calculate_btn.setStyleSheet("background-color: #4CAF50; color: white;")

        button_layout.addWidget(self.add_btn)
        button_layout.addWidget(self.edit_btn)
        button_layout.addWidget(self.delete_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.calculate_btn)

        main_layout.addLayout(button_layout)

        # Статус бар
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Готово")

    def edit_configuration(self):
        """Открытие интерактивного редактора конфигурации"""
        try:
            from config_editor import ConfigEditor

            # Получаем текущую конфигурацию
            current_config = None
            if hasattr(self.fuzzy_system, 'config'):
                current_config = self.fuzzy_system.config
            else:
                from config_editor import get_default_config
                current_config = get_default_config()

            # Создаем редактор
            editor = ConfigEditor(self, current_config)
            editor.setModal(True)

            # Показываем диалог
            if editor.exec() == QDialog.DialogCode.Accepted:
                # Получаем новую конфигурацию
                new_config = editor.get_current_config()

                try:
                    # Создаем новую систему
                    from fuzzy_system import FuzzyRiskSystem
                    self.fuzzy_system = FuzzyRiskSystem(new_config)

                    self.status_bar.showMessage("Конфигурация обновлена", 3000)

                except Exception as e:
                    error_msg = f"Не удалось применить конфигурацию:\n{str(e)[:200]}"
                    QMessageBox.critical(self, "Ошибка", error_msg)

        except Exception as e:
            error_msg = f"Ошибка при открытии редактора:\n{str(e)}"
            print(f"DEBUG ERROR: {error_msg}")
            QMessageBox.critical(self, "Ошибка", error_msg)

    def create_menu_bar(self):
        """Создание меню"""
        menubar = self.menuBar()

        # Меню Файл
        file_menu = menubar.addMenu("Файл")

        export_action = QAction("Экспорт в Excel", self)
        export_action.triggered.connect(self.export_to_excel)
        file_menu.addAction(export_action)

        file_menu.addSeparator()

        exit_action = QAction("Выход", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # Меню Настройки
        settings_menu = menubar.addMenu("Настройки")

        edit_config_action = QAction("Редактировать конфигурацию...", self)
        edit_config_action.triggered.connect(self.edit_configuration)
        settings_menu.addAction(edit_config_action)

        # Новый пункт меню для справочника предприятий
        departments_action = QAction("Управление предприятиями...", self)
        departments_action.triggered.connect(self.open_departments_window)
        settings_menu.addAction(departments_action)


    def export_to_excel(self):
        """Экспорт таблицы в Excel"""
        if not hasattr(self, 'model') or not self.model:
            QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта!")
            return


        # Диалог сохранения файла
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить как Excel",
            os.path.expanduser("~/employees_export.xlsx"),
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            # Собираем данные для экспорта
            data = []
            headers = []

            # Заголовки колонок
            for col in range(self.model.columnCount()):
                headers.append(self.model.headerData(col, Qt.Orientation.Horizontal))

            # Данные
            for row in range(self.model.rowCount()):
                row_data = []
                for col in range(self.model.columnCount()):
                    index = self.model.index(row, col)
                    value = self.model.data(index)
                    row_data.append(value)
                data.append(row_data)



            wb = Workbook()
            ws = wb.active
            ws.title = "Сотрудники"

            # Заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            # Данные
            for row, row_data in enumerate(data, 2):
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='left')

            # Автоширина колонок
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = ws.column_dimensions[chr(64 + col) if col < 27 else f'A{col}']
                for row in range(1, len(data) + 2):
                    try:
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                column.width = adjusted_width

            wb.save(file_path)

            QMessageBox.information(
                self,
                "Успешно",
                f"Данные успешно экспортированы в файл:\n{file_path}"
            )
            self.status_bar.showMessage(f"Экспорт завершен: {file_path}", 5000)

            # Открываем папку с файлом
            if os.path.exists(file_path):
                os.startfile(os.path.dirname(file_path))

        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка экспорта",
                f"Не удалось экспортировать данные:\n{str(e)}"
            )

    def on_table_double_click(self, index):
        """Обработка двойного клика по таблице"""
        self.edit_selected_employee()

    def load_employees(self, employees=None):
        """Загрузка работников в таблицу"""
        try:
            if employees is None:
                employees = self.employee_manager.get_all_employees()

            if employees:
                self.model = EmployeeTableModel(employees)
                self.table_view.setModel(self.model)

                # Настройка ширины колонок
                header = self.table_view.horizontalHeader()
                header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(10, QHeaderView.ResizeMode.Stretch)
                header.setSectionResizeMode(11, QHeaderView.ResizeMode.ResizeToContents)



                self.status_bar.showMessage(f"Загружено сотрудников: {len(employees)}", 3000)
            else:
                # Если нет сотрудников, показываем пустую модель
                self.model = EmployeeTableModel([])
                self.table_view.setModel(self.model)
                self.status_bar.showMessage("Нет данных о сотрудниках", 3000)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить сотрудников: {str(e)}")

    def open_departments_window(self):
        """Открыть окно управления предприятиями"""
        from department_manager import DepartmentsWindow
        self.departments_window = DepartmentsWindow(self)
        self.departments_window.show()

    def search_employees(self):
        """Поиск работников по трем критериям"""
        # Получаем значения из полей поиска
        name_query = self.search_name_edit.text().strip()
        position_query = self.search_position_edit.text().strip()
        department_query = self.search_department_edit.text().strip()

        # Если все поля пустые, показываем всех сотрудников
        if not name_query and not position_query and not department_query:
            self.load_employees()
            return

        try:
            # Получаем всех сотрудников
            all_employees = self.employee_manager.get_all_employees()
            filtered_employees = []

            for employee in all_employees:
                match = True

                # Поиск по ФИО
                if name_query:
                    name_match = False
                    # Поиск в фамилии
                    if employee.lastname and name_query.lower() in employee.lastname.lower():
                        name_match = True
                    # Поиск в имени
                    if employee.firstname and name_query.lower() in employee.firstname.lower():
                        name_match = True
                    # Поиск в отчестве
                    if employee.patronymic and name_query.lower() in employee.patronymic.lower():
                        name_match = True

                    if not name_match:
                        match = False

                # Поиск по должности
                if match and position_query:
                    if not (employee.position and position_query.lower() in employee.position.lower()):
                        match = False

                # Поиск по предприятию
                if match and department_query:
                    if not (employee.department_name and
                            department_query.lower() in employee.department_name.lower()):
                        match = False

                if match:
                    filtered_employees.append(employee)

            # Загружаем отфильтрованных сотрудников
            if filtered_employees:
                self.load_employees(filtered_employees)
                self.status_bar.showMessage(f"Найдено сотрудников: {len(filtered_employees)}", 3000)
            else:
                self.load_employees([])  # Показываем пустую таблицу
                self.status_bar.showMessage("Сотрудники не найдены", 3000)

        except Exception as e:
            print(f"Error searching employees: {e}")
            self.status_bar.showMessage("Ошибка при поиске", 3000)

    def clear_search(self):
        """Сброс всех полей поиска"""
        self.search_name_edit.clear()
        self.search_position_edit.clear()
        self.search_department_edit.clear()
        self.load_employees()  # Загружаем всех сотрудников

    def add_employee(self):
        """Добавление нового работника"""
        dialog = AddEditEmployeeDialog(self)
        if dialog.exec():
            data = dialog.get_employee_data()
            if data:
                try:
                    employee = self.employee_manager.add_employee(**data)
                    if employee:
                        self.load_employees()
                        self.status_bar.showMessage(f"Добавлен сотрудник: {employee.full_name}", 3000)
                    else:
                        QMessageBox.warning(self, "Ошибка", "Не удалось добавить сотрудника")
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Ошибка при добавлении: {str(e)}")

    def edit_selected_employee(self):
        """Редактирование выбранного работника"""
        selected = self.table_view.selectionModel().selectedRows()
        if not selected:
            QMessageBox.warning(self, "Ошибка", "Выберите работника для редактирования!")
            return

        try:
            row = selected[0].row()
            employee_id = self.model.employees[row].id
            employee = self.employee_manager.get_employee_by_id(employee_id)

            if employee:
                dialog = AddEditEmployeeDialog(self, employee)
                if dialog.exec():
                    data = dialog.get_employee_data()
                    if data:
                        success = self.employee_manager.update_employee(employee_id, **data)
                        if success:
                            self.load_employees()
                            self.status_bar.showMessage("Данные обновлены", 3000)
                        else:
                            QMessageBox.warning(self, "Ошибка", "Не удалось обновить данные")
            else:
                QMessageBox.warning(self, "Ошибка", "Сотрудник не найден")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при редактировании: {str(e)}")

    def delete_selected_employee(self):
        """Удаление выбранного работника"""
        selected = self.table_view.selectionModel().selectedRows()
        if not selected:
            QMessageBox.warning(self, "Ошибка", "Выберите работника для удаления!")
            return

        try:
            row = selected[0].row()
            employee_id = self.model.employees[row].id
            employee_name = self.model.employees[row].full_name

            # Подтверждение удаления
            reply = QMessageBox.question(
                self, "Подтверждение",
                f"Удалить работника {employee_name}?\nВся связанная информация также будет удалена.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                success = self.employee_manager.delete_employee(employee_id)
                if success:
                    self.load_employees()
                    self.status_bar.showMessage(f"Сотрудник {employee_name} удален", 3000)
                else:
                    QMessageBox.warning(self, "Ошибка", "Не удалось удалить сотрудника")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении: {str(e)}")

    def select_all_employees(self):
        """Выбрать всех сотрудников в таблице"""
        self.table_view.selectAll()
        self.status_bar.showMessage(f"Выбрано всех сотрудников", 3000)

    def deselect_all_employees(self):
        """Снять выделение со всех сотрудников"""
        self.table_view.clearSelection()
        self.status_bar.showMessage("Выделение снято", 3000)

    def calculate_risk_for_selected(self):
        """Расчет риска для выбранных сотрудников"""
        selected_rows = self.table_view.selectionModel().selectedRows()

        if not selected_rows:
            QMessageBox.warning(self, "Ошибка", "Выберите сотрудников для расчета риска!")
            return

        try:
            # Собираем выбранных сотрудников
            selected_employees = []
            for index in selected_rows:
                row = index.row()
                employee_id = self.model.employees[row].id
                employee = self.employee_manager.get_employee_by_id(employee_id)
                if employee:
                    selected_employees.append(employee)

            if selected_employees and self.fuzzy_system:
                dialog = MultiRiskCalculatorDialog(self, selected_employees, self.fuzzy_system)
                dialog.exec()
            elif not self.fuzzy_system:
                QMessageBox.warning(self, "Ошибка", "Система нечеткой логики не доступна")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при расчете риска: {str(e)}")